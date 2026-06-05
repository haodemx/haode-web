#!/usr/bin/env python3
"""
HAODE Product Control Center

Builds the single product workbook and health report from:
- latest local price table
- Firestore products collection
- website product data
- app/products.json
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
WORKSPACE = ROOT.parent
PRICE_TABLE_CANDIDATES = [
    Path("/Users/mac/Desktop/haode产品素材/HL CDMX 2026 06xlsx.xlsx"),
    WORKSPACE / "HAODE-WEBSITE" / "HAODE_PRODUCT_DATABASE.xlsx",
]
EXISTING_MASTER_CSV = ROOT / "docs" / "master-data" / "products-master.csv"
APP_PRODUCTS = ROOT / "app" / "products.json"
WEBSITE_PRODUCTS = ROOT / "data" / "products.generated.js"
OUTPUT_XLSX = ROOT / "data" / "products-master.xlsx"
HEALTH_REPORT = ROOT / "docs" / "reports" / "product-health-report.md"
FIRESTORE_EXPORT = ROOT / "data" / "firestore-products.json"
FIRESTORE_URL = (
    "https://firestore.googleapis.com/v1/projects/haode-app/databases/(default)"
    "/documents/products?key=AIzaSyDSDQVR_spJjvJxIpLa4k6tqoDoRhTfpPw&pageSize=500"
)

CATEGORY_MAP = {
    "iphone-incell": "Pantallas iPhone INCELL",
    "iphone-oled": "Pantallas iPhone OLED",
    "samsung-incell": "Pantallas Samsung INCELL",
    "samsung-oled": "Pantallas Samsung OLED",
    "productos-ai": "Productos AI",
    "gafas-inteligentes-ai": "Gafas AI",
    "maquinas-de-hidrogel": "Máquinas de Mica",
    "camaras-inteligentes": "Cámaras AI",
}

MAIN_HEADERS = [
    "SKU",
    "分类",
    "品牌",
    "型号",
    "产品名称",
    "零售价",
    "批发价",
    "图片路径",
    "视频路径",
    "状态",
]


@dataclass
class SourceProduct:
    sku: str
    category: str = ""
    brand: str = ""
    model: str = ""
    name: str = ""
    public_price: str = ""
    wholesale_price: str = ""
    image_path: str = ""
    video_path: str = ""
    status: str = "Activo"
    source: str = ""


@dataclass
class ProductRecord:
    sku: str
    category: str = ""
    brand: str = ""
    model: str = ""
    name: str = ""
    public_price: str = ""
    wholesale_price: str = ""
    image_path: str = ""
    video_path: str = ""
    status: str = "Activo"
    sources: dict[str, SourceProduct] = field(default_factory=dict)


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def parse_money(value: Any) -> str:
    text = normalize_text(value)
    if not text or text.lower() in {"consultar", "nan", "none"}:
        return ""
    number = re.sub(r"[^0-9.]", "", text)
    if not number:
        return ""
    parsed = float(number)
    return str(int(parsed)) if parsed.is_integer() else str(parsed)


def infer_brand(category: str, model: str, name: str) -> str:
    text = f"{category} {model} {name}".lower()
    if "iphone" in text:
        return "iPhone"
    if "samsung" in text:
        return "Samsung"
    if "haode" in text:
        return "HAODE"
    if "aimb" in text:
        return "AIMB"
    if "mica" in text:
        return "HAODE"
    if "gafa" in text or "ai" in text:
        return "HAODE"
    return ""


def public_path_to_file(value: str) -> Path | None:
    if not value:
        return None
    path = value
    if path.startswith("/haode-web/"):
        return ROOT / path.removeprefix("/haode-web/")
    if path.startswith("assets/"):
        return ROOT / path
    if path.startswith("/Users/"):
        return Path(path)
    return ROOT / path.lstrip("/")


def path_exists(value: str) -> bool:
    file_path = public_path_to_file(value)
    return bool(file_path and file_path.exists())


def make_product(
    source: str,
    sku: str,
    category: str = "",
    brand: str = "",
    model: str = "",
    name: str = "",
    public_price: Any = "",
    wholesale_price: Any = "",
    image_path: str = "",
    video_path: str = "",
    status: str = "Activo",
) -> SourceProduct:
    category = normalize_text(category)
    model = normalize_text(model)
    name = normalize_text(name)
    brand = normalize_text(brand) or infer_brand(category, model, name)
    return SourceProduct(
        sku=normalize_key(sku) or normalize_key(f"{category}-{brand}-{model}-{name}"),
        category=category,
        brand=brand,
        model=model,
        name=name,
        public_price=parse_money(public_price),
        wholesale_price=parse_money(wholesale_price),
        image_path=normalize_text(image_path),
        video_path=normalize_text(video_path),
        status=normalize_text(status) or "Activo",
        source=source,
    )


def load_existing_master() -> list[SourceProduct]:
    if not EXISTING_MASTER_CSV.exists():
        return []
    with EXISTING_MASTER_CSV.open(newline="", encoding="utf-8") as handle:
        rows = csv.DictReader(handle)
        return [
            make_product(
                "existing-master",
                row.get("id") or row.get("SKU"),
                row.get("categoria"),
                "",
                row.get("modelo"),
                row.get("producto_nombre"),
                row.get("precio_publico"),
                row.get("precio_mayoreo"),
                row.get("imagen_path"),
                row.get("video_path"),
                row.get("estado") or "Activo",
            )
            for row in rows
        ]


def current_price_table() -> Path | None:
    return next((path for path in PRICE_TABLE_CANDIDATES if path.exists()), None)


def price_category(model: str, description: str) -> str:
    text = f"{model} {description}".upper()
    if any(token in text for token in ("IPHONE", " PRO", "PROMAX", "MINI", "PLUS")) or re.match(r"^(X|XS|XR|1[1-7])", text):
        if "OLED" in text:
            return "Pantallas iPhone OLED"
        return "Pantallas iPhone INCELL"
    if "SAMSUNG" in text or re.match(r"^(S|NOTE|A[0-9])", text):
        if "OLED" in text or "AMOLED" in text:
            return "Pantallas Samsung OLED"
        return "Pantallas Samsung INCELL"
    if "MICA" in text:
        return "Micas"
    return "Refacciones"


def price_product_name(category: str, model: str) -> str:
    model = normalize_text(model)
    if "iPhone" in category:
        return f"Pantalla para iPhone {model}"
    if "Samsung" in category:
        return f"Pantalla para Samsung {model}"
    return model


def load_database_style_price_table(path: Path) -> list[SourceProduct]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["数据库"] if "数据库" in workbook.sheetnames else workbook.worksheets[0]
    rows = []
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        name, model, category, image_path, _image_count, video_count, public_price, wholesale_price, status = (list(raw) + [""] * 9)[:9]
        if not normalize_text(name) and not normalize_text(model):
            continue
        video_path = "" if not video_count else "素材库已有视频"
        rows.append(
            make_product(
                "price-table",
                "",
                category,
                "",
                model,
                name,
                public_price,
                wholesale_price,
                image_path,
                video_path,
                status or "正常",
            )
        )
    return rows


def load_hl_price_table(path: Path) -> list[SourceProduct]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows = []
    for raw in sheet.iter_rows(min_row=1, values_only=True):
        cells = list(raw) + [""] * 18
        if normalize_text(cells[0]).lower() == "item no.":
            continue
        item_no = cells[0]
        model = normalize_text(cells[3])
        description = normalize_text(cells[4])
        public_price = cells[5]
        wholesale_price = cells[6]
        if not isinstance(item_no, (int, float)) or not model or not description:
            continue
        if not parse_money(public_price) and not parse_money(wholesale_price):
            continue
        category = price_category(model, description)
        name = price_product_name(category, model)
        rows.append(
            make_product(
                "price-table",
                "",
                category,
                "",
                model,
                name,
                public_price,
                wholesale_price,
                "",
                "",
                "正常",
            )
        )
    return rows


def load_price_table() -> list[SourceProduct]:
    price_table = current_price_table()
    if not price_table:
        return []
    workbook = load_workbook(price_table, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if "数据库" in sheet_names:
        return load_database_style_price_table(price_table)
    return load_hl_price_table(price_table)


def load_app_products() -> list[SourceProduct]:
    if not APP_PRODUCTS.exists():
        return []
    data = json.loads(APP_PRODUCTS.read_text(encoding="utf-8"))
    return [
        make_product(
            "app",
            item.get("id"),
            item.get("categoria"),
            "",
            item.get("modelo"),
            item.get("nombre"),
            item.get("precioPublico"),
            item.get("precioMayoreo"),
            item.get("imagen"),
            "",
            "Activo" if item.get("activo", True) else "Inactivo",
        )
        for item in data
    ]


def load_website_products() -> list[SourceProduct]:
    if not WEBSITE_PRODUCTS.exists():
        return []
    text = WEBSITE_PRODUCTS.read_text(encoding="utf-8")
    start = text.find("[")
    end = text.rfind("]")
    if start < 0 or end < start:
        return []
    data = json.loads(text[start : end + 1])
    rows = []
    for item in data:
        prices = item.get("prices") or []
        rows.append(
            make_product(
                "website",
                item.get("id"),
                CATEGORY_MAP.get(item.get("category"), item.get("category")),
                item.get("brand"),
                " ".join(part for part in [item.get("model"), item.get("quality")] if part),
                item.get("name"),
                prices[0].get("price") if len(prices) > 0 else "",
                prices[1].get("price") if len(prices) > 1 else "",
                (item.get("images") or [""])[0],
                (item.get("videos") or [""])[0],
                "Activo",
            )
        )
    return rows


def firestore_value(value: dict[str, Any]) -> Any:
    for key in ("stringValue", "integerValue", "doubleValue", "booleanValue", "timestampValue"):
        if key in value:
            return value[key]
    return ""


def load_firestore_products(fetch_live: bool = True) -> list[SourceProduct]:
    payload = None
    if FIRESTORE_EXPORT.exists():
        payload = json.loads(FIRESTORE_EXPORT.read_text(encoding="utf-8"))
    elif fetch_live:
        with urllib.request.urlopen(FIRESTORE_URL, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
    if not payload:
        return []
    rows = []
    for doc in payload.get("documents", []):
        fields = {key: firestore_value(value) for key, value in (doc.get("fields") or {}).items()}
        rows.append(
            make_product(
                "firestore",
                fields.get("id") or doc.get("name", "").split("/")[-1],
                fields.get("categoria"),
                "",
                fields.get("modelo"),
                fields.get("nombre"),
                fields.get("precioPublico"),
                fields.get("precioMayoreo"),
                fields.get("imagen"),
                "",
                "Activo" if fields.get("activo", True) else "Inactivo",
            )
        )
    return rows


def index_online_products(products: list[SourceProduct]) -> dict[str, str]:
    index = {}
    for product in products:
        for key in {
            normalize_key(product.sku),
            normalize_key(product.model),
            normalize_key(f"{product.category}-{product.model}"),
            normalize_key(f"{product.name}-{product.model}"),
        }:
            if key:
                index.setdefault(key, product.sku)
    return index


def choose_sku(product: SourceProduct, index: dict[str, str]) -> str:
    for key in [
        normalize_key(product.model),
        normalize_key(f"{product.category}-{product.model}"),
        normalize_key(f"{product.name}-{product.model}"),
        normalize_key(product.name),
    ]:
        if key in index:
            return index[key]
    return product.sku


def merge_products(sources: dict[str, list[SourceProduct]]) -> dict[str, ProductRecord]:
    online = sources["existing-master"] + sources["app"] + sources["website"] + sources["firestore"]
    online_index = index_online_products(online)
    products: dict[str, ProductRecord] = {}
    priority = ["existing-master", "price-table", "firestore", "app", "website"]

    for source_name in priority:
        for source_product in sources[source_name]:
            sku = choose_sku(source_product, online_index) if source_name == "price-table" else source_product.sku
            if not sku:
                continue
            record = products.setdefault(sku, ProductRecord(sku=sku))
            record.sources[source_name] = source_product

            for attr in ("category", "brand", "model", "name", "image_path", "video_path", "status"):
                if not getattr(record, attr) and getattr(source_product, attr):
                    setattr(record, attr, getattr(source_product, attr))
            if source_name == "price-table" or not record.public_price:
                if source_product.public_price:
                    record.public_price = source_product.public_price
            if source_name == "price-table" or not record.wholesale_price:
                if source_product.wholesale_price:
                    record.wholesale_price = source_product.wholesale_price

    return products


def mismatch(values: list[str]) -> bool:
    cleaned = [normalize_text(value) for value in values if normalize_text(value)]
    return len(set(cleaned)) > 1


def build_checks(products: dict[str, ProductRecord]) -> dict[str, Any]:
    source_names = ["firestore", "website", "app"]
    missing_products = []
    price_errors = []
    category_errors = []
    image_missing = []
    video_missing = []
    duplicate_groups = defaultdict(list)

    for product in products.values():
        missing = [source for source in source_names if source not in product.sources]
        if missing:
            missing_products.append((product, missing))

        public_values = [product.public_price] + [product.sources[source].public_price for source in source_names if source in product.sources]
        wholesale_values = [product.wholesale_price] + [product.sources[source].wholesale_price for source in source_names if source in product.sources]
        if mismatch(public_values) or mismatch(wholesale_values):
            price_errors.append(product)

        category_values = [product.category] + [product.sources[source].category for source in source_names if source in product.sources]
        if mismatch(category_values):
            category_errors.append(product)

        if not product.image_path or not path_exists(product.image_path):
            image_missing.append(product)
        if not product.video_path or not path_exists(product.video_path):
            video_missing.append(product)

        # A duplicate must be the same catalog identity, not just the same model.
        # Example: two different cases can both target "Estilo iPhone 17 Pro Max".
        duplicate_key = normalize_key(f"{product.category}-{product.model}-{product.name}")
        if duplicate_key:
            duplicate_groups[duplicate_key].append(product)

    duplicates = [items for items in duplicate_groups.values() if len(items) > 1]
    total = len(products)
    return {
        "total": total,
        "missing_products": missing_products,
        "duplicates": duplicates,
        "price_errors": price_errors,
        "category_errors": category_errors,
        "image_missing": image_missing,
        "video_missing": video_missing,
        "image_rate": round(((total - len(image_missing)) / total) * 100, 1) if total else 0,
        "video_rate": round(((total - len(video_missing)) / total) * 100, 1) if total else 0,
    }


def status_for(product: ProductRecord, checks: dict[str, Any]) -> str:
    issues = []
    if any(item[0].sku == product.sku for item in checks["missing_products"]):
        issues.append("缺平台")
    if any(item.sku == product.sku for item in checks["price_errors"]):
        issues.append("价格异常")
    if any(item.sku == product.sku for item in checks["category_errors"]):
        issues.append("分类异常")
    if any(item.sku == product.sku for item in checks["image_missing"]):
        issues.append("图片缺失")
    if any(item.sku == product.sku for item in checks["video_missing"]):
        issues.append("视频缺失")
    return "正常" if not issues else " / ".join(issues)


def write_workbook(products: dict[str, ProductRecord], checks: dict[str, Any], sources: dict[str, list[SourceProduct]]) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "products-master"
    ws.append(MAIN_HEADERS)
    for product in sorted(products.values(), key=lambda item: (item.category, item.brand, item.model, item.sku)):
        ws.append([
            product.sku,
            product.category,
            product.brand,
            product.model,
            product.name,
            product.public_price,
            product.wholesale_price,
            product.image_path,
            product.video_path,
            status_for(product, checks),
        ])

    checks_ws = workbook.create_sheet("daily-check")
    checks_ws.append(["检查项", "数量"])
    checks_ws.append(["产品总数", checks["total"]])
    checks_ws.append(["缺产品数量", len(checks["missing_products"])])
    checks_ws.append(["重复产品数量", sum(len(group) for group in checks["duplicates"])])
    checks_ws.append(["价格异常数量", len(checks["price_errors"])])
    checks_ws.append(["分类异常数量", len(checks["category_errors"])])
    checks_ws.append(["图片缺失数量", len(checks["image_missing"])])
    checks_ws.append(["视频缺失数量", len(checks["video_missing"])])
    checks_ws.append(["图片完整率", f"{checks['image_rate']}%"])
    checks_ws.append(["视频完整率", f"{checks['video_rate']}%"])

    exceptions_ws = workbook.create_sheet("exceptions")
    exceptions_ws.append(["类型", "SKU", "产品名称", "型号", "详情"])
    for product, missing in checks["missing_products"]:
        exceptions_ws.append(["缺产品", product.sku, product.name, product.model, ", ".join(missing)])
    for group in checks["duplicates"]:
        detail = ", ".join(product.sku for product in group)
        for product in group:
            exceptions_ws.append(["重复产品", product.sku, product.name, product.model, detail])
    for label, key in [("价格错误", "price_errors"), ("分类错误", "category_errors"), ("图片缺失", "image_missing"), ("视频缺失", "video_missing")]:
        for product in checks[key]:
            exceptions_ws.append([label, product.sku, product.name, product.model, ""])

    sources_ws = workbook.create_sheet("sources")
    sources_ws.append(["来源", "产品数"])
    for source_name, rows in sources.items():
        sources_ws.append([source_name, len(rows)])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = width

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_XLSX)


def list_rows(items: list[Any], limit: int = 30) -> str:
    if not items:
        return "- 无"
    lines = []
    for item in items[:limit]:
        product = item[0] if isinstance(item, tuple) else item
        detail = f" | 缺: {', '.join(item[1])}" if isinstance(item, tuple) else ""
        lines.append(f"- {product.sku} | {product.name} | {product.model}{detail}")
    if len(items) > limit:
        lines.append(f"- 其余 {len(items) - limit} 条见 `data/products-master.xlsx` 的 exceptions 工作表")
    return "\n".join(lines)


def write_health_report(products: dict[str, ProductRecord], checks: dict[str, Any], sources: dict[str, list[SourceProduct]]) -> None:
    duplicate_count = sum(len(group) for group in checks["duplicates"])
    report = f"""# HAODE Product Control Center 健康报告

生成日期：{date.today().isoformat()}

## 数据来源

| 来源 | 产品数 |
| --- | ---: |
| products-master 当前 CSV | {len(sources['existing-master'])} |
| 当前最新价格表 | {len(sources['price-table'])} |
| Firestore | {len(sources['firestore'])} |
| 网站 | {len(sources['website'])} |
| App | {len(sources['app'])} |

## 核心指标

| 指标 | 当前值 |
| --- | ---: |
| 产品总数 | {checks['total']} |
| 缺产品数量 | {len(checks['missing_products'])} |
| 重复产品数量 | {duplicate_count} |
| 价格异常数量 | {len(checks['price_errors'])} |
| 分类异常数量 | {len(checks['category_errors'])} |
| 图片完整率 | {checks['image_rate']}% |
| 视频完整率 | {checks['video_rate']}% |

## 每日自动比对范围

- `data/products-master.xlsx`
- Firestore `products`
- 网站 `data/products.generated.js`
- App `app/products.json`

## 自动验证规则

以后任何产品修改完成后，必须运行：

```bash
npm run product-control-center
npm run product-validate
```

验证必须覆盖：

- Firestore 是否存在该产品
- 网站是否存在该产品
- App 是否存在该产品
- 三方价格是否一致
- 三方分类是否一致
- 图片路径是否存在
- 视频路径是否存在或明确标记缺素材

## 缺产品

{list_rows(checks['missing_products'])}

## 价格异常

{list_rows(checks['price_errors'])}

## 分类异常

{list_rows(checks['category_errors'])}

## 图片缺失

{list_rows(checks['image_missing'])}

## 视频缺失

{list_rows(checks['video_missing'])}
"""
    HEALTH_REPORT.parent.mkdir(parents=True, exist_ok=True)
    HEALTH_REPORT.write_text(report, encoding="utf-8")


def validate(checks: dict[str, Any]) -> int:
    failures = [
        len(checks["missing_products"]),
        len(checks["price_errors"]),
        len(checks["category_errors"]),
        len(checks["image_missing"]),
    ]
    return 1 if any(failures) else 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--validate-only", action="store_true")
    parser.add_argument("--no-firestore-live", action="store_true")
    args = parser.parse_args()

    sources = {
        "existing-master": load_existing_master(),
        "price-table": load_price_table(),
        "firestore": load_firestore_products(fetch_live=not args.no_firestore_live),
        "website": load_website_products(),
        "app": load_app_products(),
    }
    products = merge_products(sources)
    checks = build_checks(products)

    if not args.validate_only:
        write_workbook(products, checks, sources)
        write_health_report(products, checks, sources)

    summary = {
        "products": checks["total"],
        "missingProducts": len(checks["missing_products"]),
        "duplicateProducts": sum(len(group) for group in checks["duplicates"]),
        "priceErrors": len(checks["price_errors"]),
        "categoryErrors": len(checks["category_errors"]),
        "imageRate": checks["image_rate"],
        "videoRate": checks["video_rate"],
        "output": str(OUTPUT_XLSX.relative_to(ROOT)),
        "report": str(HEALTH_REPORT.relative_to(ROOT)),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return validate(checks) if args.validate_only else 0


if __name__ == "__main__":
    sys.exit(main())
